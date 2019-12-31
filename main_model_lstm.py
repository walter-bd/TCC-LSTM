#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jul 9 16:55:30 2018
@author: walter
"""
from Modules.pre_processing_data import (feature_scaling,
                                         scaling_test)
import keras.backend as K
import pandas
import matplotlib.pyplot as plt
import copy
from keras.models import Sequential
from keras.layers import Dense, LeakyReLU
from keras.layers import LSTM, Dropout, Flatten
from keras.optimizers import Adam, SGD
from keras import regularizers
from keras.utils import plot_model
import openpyxl
from math import pi
import numpy as np
import tensorflow as tf
from scipy.stats import iqr
from numpy.random import seed
from tensorflow import set_random_seed
from keras.utils import plot_model
from random import randint
from itertools import zip_longest
from scipy.ndimage.interpolation import shift
from keras import metrics

# Function to separate an array in groups of n elements 
# https://stackoverflow.com/questions/3992735/python-generator-that-groups-another-iterable-into-groups-of-n/3992765#3992765
def grouper(n, iterable, fillvalue=None):
    "grouper(3, 'ABCDEFG', 'x') --> ABC DEF Gxx"
    args = [iter(iterable)] * n
    return list(zip_longest(fillvalue=fillvalue, *args))

def train(y_train, output_dw, input_dw, n_epoch, n_batch, type_NN, nodes):
    # fix random seed for reproducibility
    seed(random_number)
    set_random_seed(random_number)
    # Shuffle data is putted on because we want the model to learn on each sequence independently. 
    shuf = True
    stateful_ = False
    x_train_input=[]
    y_train_output=[]
    # Defines the quantity of elements on each array to feed the LSTM model, in this case each X represents a scalar element not an array.
    group_n= 1
    model = Sequential()
    # Check if the training is set to an LSTM model or an MLP model 
    if (type_NN is "LSTM"):
        # Orders the input and output of the model in arrays of (group_n, T), where group_n represents the dimension of the array of X's and T the maximun amount of steps that is feeded on the LSTM model 
        # Example group_n = 2 input_dw = 3 then there is X1 = (t1, t2), X2=(t2, t3), X3=(t3, t4) and X = [X1, X2, X3]
        # Example output_dw = 3 then Y = (Y1, Y2, Y3)
        for i in range(0,len(y_train)-input_dw-output_dw + 1):
            x_train_input.append(grouper(group_n,y_train[i:i+input_dw]))
        for i in range(input_dw, len(y_train)-output_dw + 1):
            y_train_output.append(y_train[i:i+output_dw])
        y_train_output = np.array(y_train_output)
        x_train_input = np.array(x_train_input)
        #Uses Adam optimization algorithm
        optim = Adam(lr=0.0009, decay=0.5e-6)
        x_train_input = x_train_input.reshape(x_train_input.shape[0], x_train_input.shape[1],x_train_input.shape[2])
        #Add and define properties of the LSTM model  and Flattened Up
        model.add(LSTM(nodes, batch_input_shape=(n_batch, x_train_input.shape[1],x_train_input.shape[2]), kernel_regularizer=regularizers.l1_l2(0.5e-6), activation='tanh', stateful=stateful_, return_sequences=True))
        model.add(Flatten())
    # Check if the training is set to MLP model
    elif (type_NN is "MLP"): 
        # In this case grouper is not used and is always assumed that group_n = 1
        for i in range(0,len(y_train)-input_dw-output_dw + 1):
            x_train_input.append(y_train[i:i+input_dw])
        for i in range(input_dw, len(y_train)-output_dw + 1):
            y_train_output.append(y_train[i:i+output_dw])
        y_train_output = np.array(y_train_output)
        x_train_input = np.array(x_train_input)
        #Uses Adam optimization algorithm
        optim = Adam(lr=0.00095, decay=0.5e-6)
        shuf = True
        #Add and define properties of MLP model
        model.add(Dense(nodes, activation='tanh', input_dim=input_dw, kernel_regulawrizer=regularizers.l1_l2(0.5e-6)))
    #Creates a fully connected final layer output for the model
    model.add(Dense(output_dw, activation='linear'))
    model.compile(loss='mean_squared_error',
                      optimizer=optim, metrics=["mae"])
    #Variables to save the losses in the training 
    history = {}
    history["loss"]=[]
    history["val_loss"]=[]
    k = 0
    p_loss = 100
    m_loss = 100
    #Variable to store the model weights 
    last_model = model.get_weights()
    #Creates the input for the validation step and the test step and normalize it with min-max
    x_prediction_input = grouper(group_n,list(y_train[-input_dw:]))
    scaled_y_test = np.divide(np.array(y_test)-scaled_y_train[2], scaled_y_train[1]-scaled_y_train[2])
    scaled_val = np.array(scaled_y_test[-30:]).reshape(1,30)
    x_last_value = grouper(group_n,scaled_y_test[-input_dw:])
    # Reshape validation and validation input values for each model 
    if (type_NN is "LSTM"):
        x_prediction_input = np.array(x_prediction_input)
        print(x_prediction_input.shape)
        x_prediction_input = x_prediction_input.reshape(1, x_prediction_input.shape[0],x_prediction_input.shape[1])
        x_last_value = np.array(x_last_value)
        x_last_value = x_last_value.reshape(1,x_last_value.shape[0],x_last_value.shape[1])
    else:
        x_prediction_input = np.array(x_prediction_input).reshape(1,input_dw)
        x_last_value = np.array(x_last_value).reshape(1,input_dw)
    # Star of the training of the model    
    if (type_NN is "LSTM" or type_NN is "MLP"):
        print("n:"+ str(nodes) + ", i:" + str(input_dw) + ", o:" + str(output_dw))
        for i in range(n_epoch):
            history_now = model.fit(x_train_input, y_train_output,
                                    epochs=1, batch_size=n_batch,
                                    validation_data=(x_prediction_input, scaled_val), shuffle=shuf, verbose=False)
            history["loss"].append(history_now.history['loss'][-1])
            history["val_loss"].append(history_now.history['val_loss'][-1])
            # Early stopping based on training error
            if history_now.history['loss'][-1] > 0.99*p_loss:
                k+=1
                print("k:"+str(k))
            else:
                p_loss = history_now.history['loss'][-1]
                k=0
            #model.reset_states() supress this if stateful is true and epoch is greater than 1
            if np.isnan(history_now.history['loss'][-1]):
                print("nan")
                model.set_weights(last_model)
                break
            else:
                #New metric that takes in account the validation error and training_error to choose the best model
                aux = np.abs(history_now.history['mean_absolute_error'][-1]*history_now.history['val_mean_absolute_error'][-1])
                if m_loss > aux:
                    last_model = model.get_weights()
                    m_loss = aux
                    mod_loss = history_now.history['mean_absolute_error'][-1]
                    mod_val = history_now.history['val_mean_absolute_error'][-1]
                    print(history_now.history['mean_absolute_error'][-1])
                    print(history_now.history['val_mean_absolute_error'][-1])
            # If in 10 iteration the error does not improve 1% then stop the training
            if k>10:
                break 
            print(i)
        #Set the weights of the best model and saves the error history 
        model.set_weights(last_model)
        history["loss"].append(mod_loss)
        history["val_loss"].append(mod_val)
    return model, history, x_prediction_input, x_train_input, y_train_output, x_last_value

#Function to remove outliers based on 1.5 * iqr outlier detector
def remove_outliers(y):
    Q1 = np.percentile(y, 25)
    Q3 = np.percentile(y, 75)
    y_iqr = iqr(y)
    y = [y_ for y_ in y if (y_ > Q1 - 1.5 * y_iqr)]
    y_rem = [y_ for y_ in y if (y_ < Q3 + 1.5 * y_iqr)]
    return y_rem

#data differentiator
def differentiate_data(y):
    y_diff_l_e = y[-1]
    y_diff_f_e = y[0]
    y_diff = np.diff(y)
    return y_diff, y_diff_l_e, y_diff_f_e

def reverse_differentiate(y, y_f_e_train):
    y_rev_diff = np.concatenate(([y_f_e_train], y)).cumsum()
    return y_rev_diff

#Function to denormalize data
def reverse_scaling(y, max_y, min_y):
    y_new = y*(max_y-min_y) + min_y
    return y_new

if __name__=='__main__':
    #Establish the number of cores to use with tensorflow or if GPU will be used, LSTM training works better with CPU
    num_cores = 1
    GPU = False
    CPU = True
    random_number = int(np.random.random_sample()*5000)

    if GPU:
        num_GPU = 1
        num_CPU = 1
    if CPU:
        num_CPU = 1
        num_GPU = 0

    config = tf.ConfigProto(intra_op_parallelism_threads=num_cores,\
                            inter_op_parallelism_threads=num_cores, allow_soft_placement=True,\
                            device_count = {'CPU' : num_CPU, 'GPU' : num_GPU})
    session = tf.Session(config=config)
    K.set_session(session)
    # Name of the xlsx file where the data will be saved 
    name_xls = 'data_mode_error_prediction.xlsx'
    COL = 5
    # Read input data
    x = pandas.read_csv("Dados_Selecionados.csv", usecols=[COL], engine='python')
    # Read test data 
    y_pred_data = pandas.read_csv("test_data.csv", usecols=[COL], engine='python')
    y_pred_data = y_pred_data.dropna(axis=0, how='any').values.tolist()
    y_pred = []
    for index, i in enumerate(y_pred_data):
        y_pred.append(float(i[0].replace(",",".")))
    y_pred = np.array(y_pred)
    x = x.dropna(axis=0, how='any').values.tolist()
    set_test_size = 30
    y = []
    for index, i in enumerate(x):
        y.append(float(i[0].replace(",",".")))
    # Set different parameters for the training phase
    y_train = y
    n_epoch = 150
    n_batch = 1
    ns = "_S1"
    # Normalize data with min-max method
    scaled_y_train = {}
    scaled_y_train = feature_scaling(y_train, method='min-max')
    scaled_y_train = (scaled_y_train[0][:-set_test_size], scaled_y_train[1], scaled_y_train[2])
    # Hyperparameters to search with grid search
    input_dw_list = [5,10,15,20,25,30]
    number_of_LSTM_units = [5,10,15,20,25,30]
    output_dw_list = [30]
    # Type of NN: LSTM or MLP
    type_NN = "MLP"
    # Open workbook where data is to be saved
    workbook = openpyxl.load_workbook(name_xls)
    if ('Error_'+type_NN+ns) not in workbook.sheetnames:
        errorsheet = workbook.create_sheet('Error_'+type_NN+ns)
    else: 
        errorsheet = workbook['Error_'+type_NN+ns]
    
    if ('Validation_Prediction'+type_NN+ns) not in workbook.sheetnames:
        prvsheet = workbook.create_sheet('Pred_RealValue_'+type_NN+ns)
    else: 
        prvsheet = workbook['Pred_RealValue_'+type_NN+ns]
    
    predictions_horizont = {}
    predictions_validation = {}
    error = {}
    models = {}
    val_error = {}
    test_error = {}
    mult_error = {}
    # Start of grid search
    for input_dw in input_dw_list:
        y_test = np.array(y[-set_test_size-input_dw:])
        # Normalize test data
        scaled_y_test = np.divide(np.array(y_test)-scaled_y_train[2], scaled_y_train[1]-scaled_y_train[2])

        for output_dw in output_dw_list:
            for nodes in number_of_LSTM_units:
                model_fitted = 0
                # Train the model
                (model_fitted, history, x_input, x_train_input, y_train_output, x_last_input) = train(scaled_y_train[0], output_dw, input_dw, n_epoch, n_batch, type_NN, nodes)
                key = str(input_dw)+'_'+str(output_dw) + '_'  + str(nodes)
                predictions_horizont[key] = []
                # Save error 
                error[key] = [history['loss'], history['val_loss']]
                # Prediction of validation data with best model 
                prediction = model_fitted.predict(x_input)*(scaled_y_train[1]-scaled_y_train[2]) + scaled_y_train[2]
                # Prediction of test data with best model
                prediction_test = model_fitted.predict(x_last_input)*(scaled_y_train[1]-scaled_y_train[2]) + scaled_y_train[2]
                predictions_horizont[key] = prediction_test.tolist()[0]
                predictions_validation[key] = prediction.tolist()[0]
                # Calculate error 
                val_error[key] = [np.mean(np.abs(y_test[-set_test_size:] - prediction)).tolist()]
                test_error[key] = [np.mean(np.abs(y_pred - prediction_test)).tolist()]
                mult_error[key] = val_error[key]*np.array(error[key][0][-1])*np.array(error[key][1][-1])
    # Save all this data on the xls file 
    col = 2
    errorsheet.cell(row=1, column=1).value = 'Type'
    errorsheet.cell(row=1, column=2).value = 'Train Error'
    errorsheet.cell(row=1, column=3).value = 'Validation Error on Training'
    errorsheet.cell(row=1, column=4).value = 'Validation Error'
    errorsheet.cell(row=1, column=5).value = 'Test Error'
    errorsheet.cell(row=1, column=6).value = 'Mult Error'
    errorsheet.cell(row=1, column=7).value = str(random_number)
    
    for key in error.keys():
        row_ = 1
        errorsheet.cell(row=col, column=row_).value = key
        for item in error[key]:
            row_ +=1
            errorsheet.cell(row=col, column=row_).value = item[-1]
        col += 1
    col = 2
    for key in val_error.keys():
        row_ = 3
        for item in val_error[key]:
            row_ +=1
            errorsheet.cell(row=col, column=row_).value = item
        col += 1        
    col = 2
    for key in test_error.keys():
        row_ = 4
        for item in test_error[key]:
            row_ +=1
            errorsheet.cell(row=col, column=row_).value = item
        col += 1 
    col = 2
    for key in mult_error.keys():
        row_ = 5
        for item in mult_error[key]:
            row_ +=1
            errorsheet.cell(row=col, column=row_).value = item
        col += 1 
    
    col = 1
    for key in predictions_validation.keys():
        row_ = 1
        prvsheet.cell(row=row_, column=col).value = key + "_real"
        for item in y[-set_test_size:]:
            row_ +=1
            prvsheet.cell(row=row_, column=col).value = item
        row_ = 1
        prvsheet.cell(row=row_, column=col+1).value = key + "_val"
        for item in predictions_validation[key]:
            row_ +=1
            prvsheet.cell(row=row_, column=col+1).value = item
        row_ = 1
        prvsheet.cell(row=row_, column=col+2).value = key + "_test"
        for item in predictions_horizont[key]:
            row_ +=1
            prvsheet.cell(row=row_, column=col+2).value = item

        col += 3
    workbook.save(name_xls)  